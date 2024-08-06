from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

ROOT = Path(__file__).parent
XL = ROOT / 'FilesExcel' / 'AulaXL.xlsx'

# Carregar arquivo excel
book = load_workbook(XL)


sheetname = 'Planilha'
sheet: Worksheet = book[sheetname]

for row in sheet.iter_rows(min_row=2):
    for col in row:
        print(col.value)
        if col.value == 'Matheus':
            sheet.cell(col.row, 3, 10)

book.save(XL)
