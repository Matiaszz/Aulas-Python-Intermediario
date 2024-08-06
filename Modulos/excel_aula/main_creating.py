from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from pathlib import Path

ROOT = Path(__file__).parent
XL = ROOT / 'FilesExcel' / 'AulaXL.xlsx'

book = Workbook()
sheetname = 'Planilha'
# Cria uma nova area de planilhas com o nome dado
book.create_sheet(sheetname, 0)  # 0: move a planihla criada para o indice (0)


# sheet: Worksheet = book.active  # type: ignore | Seleciona a planilha padrão
sheet: Worksheet = book[sheetname]

# Remover planilha

book.remove(book['Sheet'])
sheet.cell(1, 1, 'NOME')
sheet.cell(1, 2, 'IDADE')
sheet.cell(1, 3, 'NOTA')


students = [
    ['Allan', 17, 8.5],
    ['Davi', 18, 8],
    ['Matheus', 17, 9.99],
    ['Lara', 18, 0.5]
]

# for r, student in enumerate(students, start=2):  # r = row
#     for c, value in enumerate(student, start=1):  # c = column
#         sheet.cell(r, c, value)

# faz a mesma coisa do for de cima
for student in students:
    sheet.append(student)

book.save(XL)
